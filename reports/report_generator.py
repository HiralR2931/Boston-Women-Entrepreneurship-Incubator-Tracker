"""
Exportable report generation, run on demand from live data.

Produces:
  1. A PDF board/stakeholder report (reportlab) with KPI summary,
     top performers tables, and embedded matplotlib charts.
  2. An Excel workbook (openpyxl) with one sheet per analytical view,
     native Excel charts, and light conditional formatting.

Both are callable from the Streamlit "Reports" page so users can generate
and download fresh reports on demand.
"""
import os
import sys
import io
import datetime as dt

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import seaborn as sns
import pandas as pd

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from analytics import queries as q

sns.set_theme(style="whitegrid")
BRAND_COLOR = "#7A2E8E"     # deep purple - distinct incubator brand color
ACCENT_COLOR = "#F2A541"    # warm amber accent


# ---------------------------------------------------------------------
# Chart helpers (return PNG bytes buffers, used by both PDF and, for
# quick previews, the Streamlit app)
# ---------------------------------------------------------------------

def _fig_to_buf(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=160, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf


def chart_funding_trend():
    df = q.funding_trend_over_time()
    fig, ax = plt.subplots(figsize=(7, 3.2))
    ax.bar(df["period"], df["total_amount"] / 1e6, color=BRAND_COLOR)
    ax.set_ylabel("Total Raised ($M)")
    ax.set_title("Quarterly Funding Raised Across Portfolio")
    ax.tick_params(axis="x", rotation=60, labelsize=7)
    fig.tight_layout()
    return _fig_to_buf(fig)


def chart_top_industries():
    df = q.high_growth_industries().head(8)
    fig, ax = plt.subplots(figsize=(7, 3.2))
    sns.barplot(data=df, y="industry", x="total_funding", ax=ax, color=ACCENT_COLOR)
    ax.set_xlabel("Total Funding ($)")
    ax.set_ylabel("")
    ax.set_title("Total Funding by Industry")
    fig.tight_layout()
    return _fig_to_buf(fig)


def chart_stage_distribution():
    df = q.load_startups()["stage"].value_counts().reset_index()
    df.columns = ["stage", "count"]
    fig, ax = plt.subplots(figsize=(5, 3.2))
    ax.pie(df["count"], labels=df["stage"], autopct="%1.0f%%",
           colors=sns.color_palette("PuOr", len(df)))
    ax.set_title("Startups by Funding Stage")
    fig.tight_layout()
    return _fig_to_buf(fig)


def chart_health_score_top10():
    df = q.startup_health_score().head(10)
    fig, ax = plt.subplots(figsize=(7, 3.5))
    sns.barplot(data=df, y="name", x="health_score", ax=ax, color=BRAND_COLOR)
    ax.set_xlabel("Composite Health Score (0-100)")
    ax.set_ylabel("")
    ax.set_title("Top 10 Startups by Health Score")
    fig.tight_layout()
    return _fig_to_buf(fig)


# ---------------------------------------------------------------------
# PDF REPORT
# ---------------------------------------------------------------------

def generate_pdf_report(output_path):
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleBrand", parent=styles["Title"], textColor=colors.HexColor(BRAND_COLOR)
    )
    h2 = ParagraphStyle("H2Brand", parent=styles["Heading2"], textColor=colors.HexColor(BRAND_COLOR))
    body = styles["BodyText"]

    doc = SimpleDocTemplate(output_path, pagesize=letter,
                             topMargin=0.6 * inch, bottomMargin=0.6 * inch)
    elements = []

    kpis = q.kpi_summary()
    elements.append(Paragraph("Boston Women Entrepreneurship Incubator", title_style))
    elements.append(Paragraph("Quarterly Program Report", styles["Heading3"]))
    elements.append(Paragraph(dt.date.today().strftime("Generated on %B %d, %Y"), body))
    elements.append(Spacer(1, 16))

    # KPI table
    kpi_data = [
        ["Active Startups", str(kpis["active_startups"]), "Total Startups Tracked", str(kpis["total_startups"])],
        ["Total Capital Raised", f"${kpis['total_funding_raised']:,.0f}",
         "Funding Rounds Logged", str(kpis["total_funding_rounds"])],
        ["Mentors in Network", str(kpis["total_mentors"]), "Mentorship Sessions Logged", str(kpis["total_mentorship_sessions"])],
        ["Avg. Session Rating", f"{kpis['avg_session_rating']}/5", "Investors Engaged", str(kpis["total_investors"])],
    ]
    kpi_table = Table(kpi_data, colWidths=[1.9 * inch, 1.6 * inch, 2.0 * inch, 1.0 * inch])
    kpi_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor(BRAND_COLOR)),
        ("TEXTCOLOR", (2, 0), (2, -1), colors.HexColor(BRAND_COLOR)),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("LINEBELOW", (0, 0), (-1, -1), 0.4, colors.lightgrey),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 20))

    # Charts
    for title, chart_fn in [
        ("Funding Trend", chart_funding_trend),
        ("Funding by Industry", chart_top_industries),
        ("Startups by Stage", chart_stage_distribution),
        ("Top 10 Startups by Health Score", chart_health_score_top10),
    ]:
        elements.append(Paragraph(title, h2))
        buf = chart_fn()
        elements.append(Image(buf, width=6.2 * inch, height=6.2 * inch * 0.45))
        elements.append(Spacer(1, 10))

    elements.append(PageBreak())

    # Top funded startups table
    elements.append(Paragraph("Top 10 Funded Startups", h2))
    top_funded = q.top_funded_startups(10)
    table_data = [["#", "Startup", "Industry", "Stage", "Raised", "Rounds"]]
    for i, r in enumerate(top_funded.itertuples(), 1):
        table_data.append([i, r.name, r.industry, r.stage, f"${r.total_raised:,.0f}", r.rounds])
    elements.append(_styled_table(table_data, [0.4, 1.9, 1.3, 0.9, 1.1, 0.6]))
    elements.append(Spacer(1, 16))

    # Top mentors table
    elements.append(Paragraph("Top Mentors by Rating", h2))
    top_mentors = q.top_mentors_by_rating(10)
    table_data = [["#", "Mentor", "Expertise", "Sessions", "Avg Rating"]]
    for i, r in enumerate(top_mentors.itertuples(), 1):
        table_data.append([i, r.name, r.expertise, r.sessions_logged, r.avg_logged_rating])
    elements.append(_styled_table(table_data, [0.4, 1.8, 1.7, 1.0, 1.2]))
    elements.append(Spacer(1, 16))

    # Event ROI table
    elements.append(Paragraph("Event ROI Snapshot (top 8 by follow-on funding)", h2))
    roi = q.event_roi().head(8)
    table_data = [["Event", "Type", "Attendees", "Cost/Attendee", "Raised within 60d"]]
    for r in roi.itertuples():
        table_data.append([
            r.name[:32], r.event_type, r.attendees,
            f"${r.cost_per_attendee:,.0f}" if r.cost_per_attendee else "n/a",
            f"${r.follow_on_funding_60d:,.0f}",
        ])
    elements.append(_styled_table(table_data, [1.9, 1.2, 0.9, 1.1, 1.4]))

    elements.append(Spacer(1, 20))
    elements.append(Paragraph(
        "Note: Health scores and event ROI figures are heuristic signals intended "
        "to help program staff prioritize attention -- not audited financial figures.",
        ParagraphStyle("footnote", parent=body, fontSize=7, textColor=colors.grey),
    ))

    doc.build(elements)
    return output_path


def _styled_table(data, col_widths_in):
    t = Table(data, colWidths=[w * inch for w in col_widths_in])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(BRAND_COLOR)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5EEF8")]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    return t


# ---------------------------------------------------------------------
# EXCEL WORKBOOK
# ---------------------------------------------------------------------

def _write_df_sheet(wb, sheet_name, df, header_fill=BRAND_COLOR):
    ws = wb.create_sheet(sheet_name)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    header_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill(start_color=header_fill.replace("#", ""), end_color=header_fill.replace("#", ""), fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 40)
    return ws


def generate_excel_report(output_path):
    wb = Workbook()
    wb.remove(wb.active)

    kpis = q.kpi_summary()
    ws = wb.create_sheet("Summary")
    ws.append(["Metric", "Value"])
    for k, v in kpis.items():
        ws.append([k.replace("_", " ").title(), v])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=BRAND_COLOR.replace("#", ""), fill_type="solid")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18

    top_funded = q.top_funded_startups(20)
    ws2 = _write_df_sheet(wb, "Top Funded Startups", top_funded)
    chart = BarChart()
    chart.title = "Top Funded Startups"
    chart.y_axis.title = "Total Raised ($)"
    data_ref = Reference(ws2, min_col=5, min_row=1, max_row=min(11, len(top_funded) + 1))
    cats_ref = Reference(ws2, min_col=2, min_row=2, max_row=min(11, len(top_funded) + 1))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws2.add_chart(chart, "I2")

    top_mentors = q.top_mentors_by_rating(20)
    _write_df_sheet(wb, "Top Mentors", top_mentors)

    industries = q.high_growth_industries()
    ws4 = _write_df_sheet(wb, "Industries", industries)
    chart2 = BarChart()
    chart2.title = "Total Funding by Industry"
    data_ref = Reference(ws4, min_col=3, min_row=1, max_row=len(industries) + 1)
    cats_ref = Reference(ws4, min_col=1, min_row=2, max_row=len(industries) + 1)
    chart2.add_data(data_ref, titles_from_data=True)
    chart2.set_categories(cats_ref)
    ws4.add_chart(chart2, "G2")

    trend = q.funding_trend_over_time()
    ws5 = _write_df_sheet(wb, "Funding Trend", trend)
    linechart = LineChart()
    linechart.title = "Quarterly Funding Trend"
    data_ref = Reference(ws5, min_col=2, min_row=1, max_row=len(trend) + 1)
    cats_ref = Reference(ws5, min_col=1, min_row=2, max_row=len(trend) + 1)
    linechart.add_data(data_ref, titles_from_data=True)
    linechart.set_categories(cats_ref)
    ws5.add_chart(linechart, "E2")

    _write_df_sheet(wb, "Investor Engagement", q.investor_engagement())
    _write_df_sheet(wb, "Investor Diversity", q.investor_portfolio_diversity())
    _write_df_sheet(wb, "Startup Health Scores", q.startup_health_score())
    _write_df_sheet(wb, "Mentor Load Balance", q.mentor_load_balance())
    _write_df_sheet(wb, "Event ROI", q.event_roi())
    _write_df_sheet(wb, "Funding by Stage", q.funding_distribution_by_stage())

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    out_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "sample_reports")
    os.makedirs(out_dir, exist_ok=True)
    pdf_path = generate_pdf_report(os.path.join(out_dir, "incubator_board_report.pdf"))
    xlsx_path = generate_excel_report(os.path.join(out_dir, "incubator_analytics_workbook.xlsx"))
    print("Generated:", pdf_path)
    print("Generated:", xlsx_path)
